# uses numpy, scipy, matplotlib packages
from pylab import *
from openpyxl import *
from mpl_toolkits.mplot3d import Axes3D

# read three-col data file x and two y values
data1 = loadtxt( "example5odea.dat", 'float' )
fig = figure()
plot(data1[:,2], data1[:,5], 'k-')
plot(data1[:,3], data1[:,6], 'b-')
plot(data1[:,4], data1[:,7], 'g-')
xlabel("x")
ylabel("y")

#savefig('N.png')
#savefig( 'example1py.pdf' )
show()
"""
fig2 = figure()
plot(data1[:,0], data1[:,2], 'b-')
title(" y(t) vs t")
xlabel( "t" )
ylabel( "y(t)" )
savefig("y.png")
fig3 = figure()
plot(data1[:,0], data1[:,3], 'g-')
title(" z(t) vs t")
xlabel( "t" )
ylabel( "z(t)" )
savefig("z.png")
show()

wb = load_workbook('4excel.xlsx')

ws = wb.active
n = 4000
for i in range(0, 64000, n):
	ws['A' + str(i/n+1)] = data1[:,0][i]
	ws['B' + str(i/n+1)] = "&"
	ws['C' + str(i/n+1)] = data1[:,1][i]
	ws['D' + str(i/n+1)] = "&"
	ws['E' + str(i/n+1)] = data1[:,2][i]
	ws['F' + str(i/n+1)] = "&"
	ws['G' + str(i/n+1)] = data1[:,3][i]
	ws['H' + str(i/n+1)] = "\\\\"

wb.save('4excelfile.xlsx')
"""